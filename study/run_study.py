"""run_study.py — Étude de marché récurrente (framework figé).

Usage (après le cycle hebdo de scraps --full) :
    scraper/.venv/Scripts/python.exe study/run_study.py

Fait, dans l'ordre :
  1. lit study/config.json (paramètres FIGÉS, versionnés) ;
  2. calcule sur Supabase : stats par khet (double médiane par condo, strate 0-1BR),
     opportunités par quartier expat, biens école+métro, tension délistées ;
  3. écrit un SNAPSHOT daté (study/snapshots/YYYY-MM-DD.json) — la mémoire longue ;
  4. génère l'étude datée (docs/etudes/etude-YYYY-MM-DD.md) avec tables d'ÉVOLUTION
     calculées sur tous les snapshots antérieurs + narratif manuel (study/context.md).

Comparabilité : ne changer config.json qu'en incrémentant config_version ;
le rapport et les snapshots portent la version pour tracer les ruptures de série.
"""
from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from glob import glob

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CFG = json.load(open(os.path.join(ROOT, "study", "config.json"), encoding="utf-8"))

for line in open(os.path.join(ROOT, "scraper", ".env"), encoding="utf-8"):
    line = line.strip()
    if line and not line.startswith("#") and "=" in line:
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip())
sys.path.insert(0, os.path.join(ROOT, "scraper"))
import psycopg  # noqa: E402

TODAY = datetime.now(timezone.utc).strftime("%Y-%m-%d")

# ───────────────────────── helpers ─────────────────────────
def thb(x):
    if x is None:
        return "—"
    if x >= 1e6:
        return f"{x / 1e6:.2f} M".replace(".", ",")
    return f"{x:,.0f}".replace(",", " ")

def median(vals):
    v = sorted(x for x in vals if x is not None and x > 0)
    if not v:
        return None
    m = len(v) // 2
    return v[m] if len(v) % 2 else (v[m - 1] + v[m]) / 2

def winsorize(vals):
    lo_p, hi_p = CFG["stats"]["winsor_pct"]
    if len(vals) < CFG["stats"]["winsor_min_n"]:
        return vals
    s = sorted(vals)
    q = lambda p: s[min(len(s) - 1, int(p / 100 * len(s)))]  # noqa: E731
    lo, hi = q(lo_p), q(hi_p)
    return [min(hi, max(lo, v)) for v in vals]

def norm_condo(name):
    if not name:
        return ""
    s = name.split(",")[0].lower()
    s = "".join(ch for ch in s if ch.isalnum() or ch == " ")
    return " ".join(s.split())

def haversine(lat1, lng1, lat2, lng2):
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))

def pt_seg_dist_m(lat, lng, a, b):
    ky = 110_570.0
    kx = 111_320.0 * math.cos(math.radians(13.75))
    px, py = lng * kx, lat * ky
    ax, ay = a[0] * kx, a[1] * ky
    bx, by = b[0] * kx, b[1] * ky
    dx, dy = bx - ax, by - ay
    if dx == dy == 0:
        return math.hypot(px - ax, py - ay)
    t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / (dx * dx + dy * dy)))
    return math.hypot(px - (ax + t * dx), py - (ay + t * dy))

def is_01(beds):
    return beds is not None and beds <= 1

# ───────────────────────── données ─────────────────────────
def fetch_all():
    """Lit `listings_sane`, PAS `listings`.

    Corrigé le 2026-08-03. L'étude appliquait ses propres bornes depuis
    `study/config.json` — une TROISIÈME définition, à côté de `lib/market-bounds.ts`
    et de la vue `listings_sane`. CLAUDE.md exige une source unique ; il y en avait
    trois, et celle de l'étude divergeait (loyer plafonné à 200 000 au lieu de
    500 000, aucun plancher, aucune borne de surface).

    ⚠ Mesuré avant de corriger, parce que l'ampleur importe autant que l'existence :
    sur 36 quartiers, UN SEUL voit un chiffre bouger, de +0,5 %. La double médiane
    par immeuble absorbe ces valeurs — c'est exactement ce pour quoi elle a été
    choisie. La divergence était donc un piège de MAINTENANCE, pas une erreur de
    publication : aucun chiffre déjà publié n'est invalidé.
    """
    dsn = os.environ["SUPABASE_DB_URL"]
    win = CFG["tension"]["delisted_window_days"]
    with psycopg.connect(dsn, connect_timeout=30) as conn:
        cur = conn.execute("""
            SELECT id, source, source_url, title, deal_type, price, area_sqm, price_per_sqm,
                   bedrooms, bathrooms, condo_name, khet, lat, lng, first_seen::text
            FROM listings_sane WHERE status='active' AND price_per_sqm > 0
              AND condo_name IS NOT NULL
        """)
        cols = [c.name for c in cur.description]
        actives = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur = conn.execute(f"""
            SELECT source, khet, condo_name, bedrooms, price, first_seen::text, delisted_at::text
            FROM listings_sane
            WHERE status='inactive' AND deal_type='sale'
              AND delisted_at >= now() - interval '{win} days'
        """)
        cols = [c.name for c in cur.description]
        delisted = [dict(zip(cols, r)) for r in cur.fetchall()]
        db_start = conn.execute("SELECT min(first_seen)::date::text FROM listings").fetchone()[0]
    for r in actives:
        for k in ("price", "area_sqm", "price_per_sqm"):
            if r[k] is not None:
                r[k] = float(r[k])
    for r in delisted:
        if r["price"] is not None:
            r["price"] = float(r["price"])
    return actives, delisted, db_start

def load_geo():
    p = json.load(open(os.path.join(ROOT, CFG["paths"]["pois_geojson"]), encoding="utf-8"))
    c = json.load(open(os.path.join(ROOT, CFG["paths"]["corridors_geojson"]), encoding="utf-8"))
    schools, metros, future = [], [], []
    for f in p["features"]:
        if f["geometry"]["type"] != "Point":
            continue
        lng, lat = f["geometry"]["coordinates"][:2]
        cat = f["properties"].get("category")
        name = f["properties"].get("name") or ""
        if cat == "school":
            schools.append((lat, lng, name))
        elif cat == "metro_station":
            metros.append((lat, lng, name))
    for f in c["features"]:
        if f["properties"].get("category") == "future_line" and f["geometry"]["type"] == "LineString":
            future.append((f["properties"].get("name") or "", f["geometry"]["coordinates"]))
    return schools, metros, future

# ───────────────────────── calculs ─────────────────────────
def sale_ok(r):
    return (r["deal_type"] == "sale"
            and CFG["stats"]["sale_min_thb"] <= (r["price"] or 0) <= CFG["stats"]["sale_max_thb"])

def rent_ok(r):
    return r["deal_type"] == "rent" and (r["price"] or 0) <= CFG["stats"]["rent_max_thb_month"]

#: Les 50 quartiers de Bangkok, lus DEPUIS LE DÉCOUPAGE lui-même
#: (public/data/bangkok-khet.geojson). Jamais recopiés : une liste tenue à la
#: main finit par diverger de la carte qu'elle est censée décrire.
def _khets_bangkok() -> set[str]:
    p = os.path.join(ROOT, "public", "data", "bangkok-khet.geojson")
    if not os.path.exists(p):
        return set()
    g = json.load(open(p, encoding="utf-8"))
    return {(f.get("properties") or {}).get("name_en")
            or (f.get("properties") or {}).get("name")
            for f in g.get("features", [])} - {None}


KHETS_BKK = _khets_bangkok()


def khet_stats(actives, strata="0-1BR"):
    """Double médiane par condo + rendement within-condo, par khet.

    PÉRIMÈTRE. On ne retient que les 50 quartiers de Bangkok. Avant le
    2026-08-03, toute chaîne présente dans `khet` produisait une ligne de
    classement — y compris des annonces situées HORS de Bangkok :

      · 27 annonces « Bang Na » à 13,6578/100,6029 — Sukhumvit 107, secteur
        Bearing, en province de Samut Prakan. Elles produisaient une ligne
        « Bang Na » à côté de « Bang Na District », visuellement identique
        puisque l'affichage retire le suffixe, mais avec un rendement différent.
      · 14 « Bang Phli », 3 « Pak Kret », 2 « Bang Sao Thong » — Samut Prakan
        et Nonthaburi, correctement nommées par leurs sources.

    Ces annonces ne sont pas fausses : elles sont hors sujet. Les compter dans un
    classement des quartiers de Bangkok crée un quartier fantôme qu'un lecteur
    ne peut pas interpréter.
    """
    cs, cr = defaultdict(lambda: defaultdict(list)), defaultdict(lambda: defaultdict(list))
    for r in actives:
        if not r["khet"] or (KHETS_BKK and r["khet"] not in KHETS_BKK):
            continue
        if strata == "0-1BR" and not is_01(r["bedrooms"]):
            continue
        key = norm_condo(r["condo_name"])
        if sale_ok(r):
            cs[r["khet"]][key].append(r["price_per_sqm"])
        elif rent_ok(r):
            cr[r["khet"]][key].append(r["price_per_sqm"])
    out = {}
    for khet in set(cs) | set(cr):
        smeds = {k: median(winsorize(v)) for k, v in cs.get(khet, {}).items()}
        rmeds = {k: median(winsorize(v)) for k, v in cr.get(khet, {}).items()}
        paired = [rmeds[k] * 12 / smeds[k] * 100
                  for k in smeds if k in rmeds and smeds[k] and rmeds[k]]
        y = median(paired) if len(paired) >= CFG["stats"]["min_paired_condos"] else None
        out[khet] = {
            "sale_condos": len(smeds), "rent_condos": len(rmeds), "paired": len(paired),
            "sale_psqm": round(median(list(smeds.values()))) if smeds else None,
            "rent_psqm": round(median(list(rmeds.values())), 1) if rmeds else None,
            "yield_wc": round(y, 2) if y else None,
            "low_sample": len(smeds) < CFG["stats"]["low_sample_condos"]
                          or len(rmeds) < CFG["stats"]["low_sample_condos"],
        }
    return out

def build_condo_index(actives):
    condo_sales, condo_rents = defaultdict(list), defaultdict(list)
    for r in actives:
        key = norm_condo(r["condo_name"])
        if not key:
            continue
        if sale_ok(r):
            condo_sales[key].append(r["price_per_sqm"])
        elif rent_ok(r):
            condo_rents[key].append(r["price_per_sqm"])
    return condo_sales, condo_rents

def enrich(r, condo_sales, condo_rents):
    o = CFG["opportunities"]
    key = norm_condo(r["condo_name"])
    sales, rents = condo_sales.get(key, []), condo_rents.get(key, [])
    others = list(sales)
    if r["price_per_sqm"] in others and len(others) >= 3:
        others.remove(r["price_per_sqm"])
    sale_med, rent_med = median(others), median(rents)
    disc = (sale_med - r["price_per_sqm"]) / sale_med * 100 if sale_med else None
    yld = rent_med * 12 / r["price_per_sqm"] * 100 if rent_med else None
    return {
        "id": r["id"], "url": r["source_url"], "source": r["source"],
        "condo": r["condo_name"].split(",")[0], "khet": r["khet"],
        "price": r["price"], "ppsqm": round(r["price_per_sqm"]),
        "beds": r["bedrooms"], "area": r["area_sqm"],
        "discount_pct": round(disc, 1) if disc is not None else None,
        "yield_pct": round(yld, 1) if yld is not None else None,
        "n_sale_condo": len(sales), "n_rent_condo": len(rents),
        "first_seen": r["first_seen"][:10],
        "flag": (disc or 0) > o["flag_discount_pct"] or (yld or 0) > o["flag_yield_pct"],
    }

def opportunities(actives, condo_sales, condo_rents):
    o = CFG["opportunities"]

    def suspect(e):
        return (e["discount_pct"] or 0) > o["suspect_discount_pct"] \
            or (e["yield_pct"] or 0) > o["suspect_yield_pct"]

    def score(e):
        d = (e["discount_pct"] or 0) if e["n_sale_condo"] >= o["min_sales_in_condo_for_discount"] else 0
        y = (e["yield_pct"] or 0) if e["n_rent_condo"] >= o["min_rents_in_condo_for_yield"] else 0
        s = min(max(d, 0), o["score_discount_cap_pct"]) \
            + max(min(y, o["score_yield_cap_pct"]) - o["score_yield_baseline_pct"], 0) * 3
        return s * 0.25 if suspect(e) else s

    A = {}
    for khet in CFG["expat_khets"]:
        cands = []
        for r in actives:
            if r["khet"] != khet or not sale_ok(r):
                continue
            e = enrich(r, condo_sales, condo_rents)
            if (e["n_sale_condo"] >= o["min_sales_in_condo_for_discount"] and (e["discount_pct"] or 0) > 0) or \
               (e["n_rent_condo"] >= o["min_rents_in_condo_for_yield"] and (e["yield_pct"] or 0) >= o["yield_floor_pct"]):
                cands.append(e)
        cands.sort(key=score, reverse=True)
        seen, top = set(), []
        for e in cands:
            k = norm_condo(e["condo"])
            if CFG["opportunities"]["one_per_condo"] and k in seen:
                continue
            seen.add(k)
            top.append(e)
            if len(top) >= o["top_n_per_khet"]:
                break
        A[khet] = top
    return A

def school_metro(actives, condo_sales, condo_rents, schools, metros, future):
    p = CFG["proximity"]
    car = set(CFG["car_dependent_khets"])
    B = []
    for r in actives:
        if not sale_ok(r) or r["khet"] in car or r["lat"] is None:
            continue
        ds, sname = min(((haversine(r["lat"], r["lng"], a, b), n) for a, b, n in schools),
                        default=(None, None))
        if ds is None or ds > p["school_max_m"]:
            continue
        dm, mname = min(((haversine(r["lat"], r["lng"], a, b), n) for a, b, n in metros),
                        default=(None, None))
        df = fname = None
        if dm is None or dm > p["metro_max_m"]:
            best = None
            for name, coords in future:
                if haversine(r["lat"], r["lng"], coords[0][1], coords[0][0]) > 25000:
                    continue
                for i in range(len(coords) - 1):
                    d = pt_seg_dist_m(r["lat"], r["lng"], coords[i], coords[i + 1])
                    if best is None or d < best:
                        best, fname = d, name
            df = best
            if df is None or df > p["future_line_max_m"]:
                continue
        e = enrich(r, condo_sales, condo_rents)
        if not e["yield_pct"] or e["n_rent_condo"] < p["min_rents_in_condo"]:
            continue
        e.update({
            "school_m": round(ds), "school": sname,
            "metro_m": round(dm) if dm is not None and dm <= p["metro_max_m"] else None,
            "metro": mname if dm is not None and dm <= p["metro_max_m"] else None,
            "future_m": round(df) if df is not None else None, "future_line": fname,
        })
        B.append(e)
    B.sort(key=lambda e: (min(e["yield_pct"], p["sort_yield_cap_pct"]) - (2 if e["flag"] else 0),
                          e["discount_pct"] or 0), reverse=True)
    return B[:p["top_n_export"]]

def tension(actives, delisted):
    t = CFG["tension"]
    sources = {d["source"] for d in delisted}
    active_sale = [r for r in actives if r["deal_type"] == "sale" and r["source"] in sources]

    def days(d):
        try:
            fs = datetime.fromisoformat(d["first_seen"]).astimezone(timezone.utc)
            da = datetime.fromisoformat(d["delisted_at"]).astimezone(timezone.utc)
            return max((da - fs).days, 0)
        except Exception:
            return None

    act_by_khet = Counter(r["khet"] for r in active_sale if r["khet"])
    by_khet = defaultdict(list)
    for d in delisted:
        if d["khet"]:
            by_khet[d["khet"]].append(d)
    khet_rows = []
    for khet, ds in by_khet.items():
        act = act_by_khet.get(khet, 0)
        khet_rows.append({
            "khet": khet, "n_delisted": len(ds), "active_sale": act,
            "delist_rate_pct": round(len(ds) / (len(ds) + act) * 100, 1) if (len(ds) + act) else None,
            "median_days_active": median([a for a in (days(d) for d in ds) if a is not None]),
            "median_price": median([d["price"] for d in ds if d["price"]]),
        })
    khet_rows.sort(key=lambda x: -(x["delist_rate_pct"] or 0))

    b1, b2, b3 = t["price_bands_thb"]

    def bband(p):
        p = p or 0
        return f"<{b1/1e6:.0f}M" if p < b1 else (
            f"{b1/1e6:.0f}-{b2/1e6:.0f}M" if p < b2 else (
                f"{b2/1e6:.0f}-{b3/1e6:.0f}M" if p < b3 else f">{b3/1e6:.0f}M"))

    def bbeds(b):
        return "studio/1BR" if (b or 0) <= 1 else ("2BR" if b == 2 else "3BR+")

    typo_del = Counter((bbeds(d["bedrooms"]), bband(d["price"])) for d in delisted)
    typo_act = Counter((bbeds(r["bedrooms"]), bband(r["price"])) for r in active_sale)
    typo_rows = []
    for key, n in typo_del.most_common():
        act = typo_act.get(key, 0)
        typo_rows.append({
            "typo": f"{key[0]} | {key[1]}", "n_delisted": n, "active": act,
            "churn_pct": round(n / (n + act) * 100, 1) if (n + act) else None,
        })
    return {"delisted_total": len(delisted), "sources": sorted(sources),
            "by_khet": khet_rows, "by_typology": typo_rows}

# ───────────────────────── données officielles (DOPA, BOT…) ─────────────────────────
def load_official():
    """Rafraîchit puis charge les données officielles (best effort — l'étude tourne sans)."""
    try:
        import fetch_official
        out = fetch_official.main()
        if out:
            return out
    except Exception as e:
        print(f"  (fetch officiel échoué : {str(e)[:80]})")
    try:
        return json.load(open(os.path.join(ROOT, "study", "official", "official-latest.json"),
                              encoding="utf-8"))
    except Exception:
        return None

def per_1000(actives, official):
    """Par khet : population DOPA + annonces actives (toutes strates) pour 1000 habitants."""
    if not official or "dopa" not in official:
        return {}
    pops = official["dopa"]["population_by_khet"]
    counts = defaultdict(lambda: {"sale": 0, "rent": 0})
    for r in actives:
        if r["khet"] and r["deal_type"] in ("sale", "rent"):
            counts[r["khet"]][r["deal_type"]] += 1
    out = {}
    for khet, pop in pops.items():
        c = counts.get(khet, {"sale": 0, "rent": 0})
        out[khet] = {
            "population": pop,
            "sale_per_1000": round(c["sale"] / pop * 1000, 2) if pop else None,
            "rent_per_1000": round(c["rent"] / pop * 1000, 2) if pop else None,
        }
    return out

# ───────────────────────── snapshot & évolution ─────────────────────────
def panier_constant_md(a, b, mini_par_khet=8):
    """Évolution à PANIER CONSTANT : uniquement les immeubles présents aux deux dates.

    C'est la mesure qui résiste à la croissance du corpus. On ne compare pas deux
    médianes de quartier calculées sur deux populations différentes — on prend
    chaque immeuble connu des deux côtés, on regarde de combien SA médiane a
    bougé, et on prend la médiane de ces variations.

    Ainsi un quartier qui gagne trente immeubles bon marché ne « baisse » pas :
    les nouveaux ne sont simplement pas comptés dans la variation. Ils le seront
    à l'édition suivante, une fois présents des deux côtés.
    """
    ca, cb = a.get("condo_medians"), b.get("condo_medians")
    if not ca or not cb:
        return ("> *Évolution à panier constant indisponible : les instantanés "
                "antérieurs au 2026-08-03 ne conservaient que le niveau quartier. "
                "Elle apparaîtra dès que deux éditions consécutives porteront les "
                "médianes par immeuble.*\n")

    communs = defaultdict(lambda: {"v": [], "l": []})
    for k in set(ca) & set(cb):
        kh = cb[k].get("khet") or ca[k].get("khet")
        if not kh:
            continue
        for champ, cle in (("vente_m2", "v"), ("loyer_m2", "l")):
            x, y = ca[k].get(champ), cb[k].get(champ)
            if x and y:
                communs[kh][cle].append(100 * (y / x - 1))

    lignes = []
    for kh, d in communs.items():
        if len(d["v"]) >= mini_par_khet:
            lignes.append((kh, len(d["v"]), median(d["v"]) if d["v"] else None,
                           len(d["l"]), median(d["l"]) if d["l"] else None))
    if not lignes:
        return ("> *Panier constant : aucun quartier n'atteint le seuil "
                f"de {mini_par_khet} immeubles présents aux deux dates.*\n")

    lignes.sort(key=lambda x: -(x[2] or 0))
    out = [f"\n**À PANIER CONSTANT — immeubles présents le {a['date']} ET le {b['date']}.**",
           "Ces variations sont insensibles à l'arrivée de nouveaux immeubles : "
           "seuls ceux déjà connus des deux côtés sont comptés. C'est le chiffre à lire "
           "quand le corpus a changé de taille.\n",
           "| Quartier | Immeubles | Δ prix/m² | Immeubles (loc.) | Δ loyer/m² |",
           "|---|---:|---:|---:|---:|"]
    for kh, nv, dv, nl, dl in lignes:
        out.append(f"| {kh.replace(' District', '')} | {nv} | "
                   f"{f'{dv:+.1f} %' if dv is not None else '—'} | {nl or '—'} | "
                   f"{f'{dl:+.1f} %' if dl is not None else '—'} |")
    return "\n".join(out) + "\n"


def medianes_par_condo(actives, strata="0-1BR"):
    """Médiane de prix et de loyer au m², PAR IMMEUBLE.

    POURQUOI ON GARDE CE NIVEAU DE DÉTAIL. Le corpus grossit à chaque scrap —
    c'est le principe même d'un observatoire, et les écarts entre éditions
    mélangent donc mouvement de prix et changement de couverture. Un avertissement
    ne règle rien : il dit que le chiffre est douteux sans en fournir un bon.

    La réponse est de comparer LE MÊME PANIER : les immeubles présents dans les
    deux éditions. Un immeuble déjà connu dont la médiane bouge, c'est du marché.
    Un immeuble qui apparaît, c'est de la couverture. Les deux sont intéressants,
    mais ce sont deux mesures, et les additionner n'en donne aucune.

    Impossible à reconstituer après coup : jusqu'au 2026-08-03 les instantanés ne
    gardaient que le niveau quartier. La série de panier constant commencera donc
    à la première édition qui suivra celle-ci.

    Coût : ~3 000 immeubles, quelques dizaines de kilo-octets par instantané,
    pour une cadence mensuelle. Négligeable au regard de ce qu'on ne peut pas
    recalculer plus tard.
    """
    ventes, loyers, khets = defaultdict(list), defaultdict(list), {}
    for r in actives:
        if not r["khet"] or (KHETS_BKK and r["khet"] not in KHETS_BKK):
            continue
        if strata == "0-1BR" and not is_01(r["bedrooms"]):
            continue
        k = norm_condo(r["condo_name"])
        if not k:
            continue
        khets.setdefault(k, r["khet"])
        if sale_ok(r):
            ventes[k].append(r["price_per_sqm"])
        elif rent_ok(r):
            loyers[k].append(r["price_per_sqm"])

    out = {}
    for k in set(ventes) | set(loyers):
        v, l = ventes.get(k, []), loyers.get(k, [])
        out[k] = {"khet": khets.get(k), "n_v": len(v), "n_l": len(l)}
        if v:
            out[k]["vente_m2"] = round(median(v))
        if l:
            out[k]["loyer_m2"] = round(median(l), 1)
    return out


def write_snapshot(actives, D01, tens, demo=None, official=None):
    snap = {
        "date": TODAY, "config_version": CFG["config_version"],
        "totals": {
            "actives": len(actives),
            "sales": sum(1 for r in actives if r["deal_type"] == "sale"),
            "rents": sum(1 for r in actives if r["deal_type"] == "rent"),
            "condos": len({norm_condo(r["condo_name"]) for r in actives}),
        },
        # tous les khets (les tables d'évolution du rapport filtrent sur key_khets_evolution,
        # les exports CSV/xlsx gardent tout)
        "khet_stats_01": {k: {kk: v[kk] for kk in ("sale_psqm", "rent_psqm", "yield_wc",
                                                   "paired", "sale_condos", "rent_condos")}
                          for k, v in D01.items()},
        "tension_by_khet": {r["khet"]: {"rate": r["delist_rate_pct"], "n": r["n_delisted"]}
                            for r in tens["by_khet"] if r["n_delisted"] >= CFG["tension"]["min_delisted_per_khet"]},
        # Médianes par IMMEUBLE — la matière du panier constant (cf. medianes_par_condo).
        "condo_medians": medianes_par_condo(actives),
        "demographics": demo or {},
        "dopa_yymm": (official or {}).get("dopa", {}).get("yymm_buddhist"),
    }
    sdir = os.path.join(ROOT, CFG["paths"]["snapshots_dir"])
    os.makedirs(sdir, exist_ok=True)
    path = os.path.join(sdir, f"{TODAY}.json")
    json.dump(snap, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    return path

def load_snapshots():
    sdir = os.path.join(ROOT, CFG["paths"]["snapshots_dir"])
    snaps = []
    for f in sorted(glob(os.path.join(sdir, "*.json"))):
        try:
            snaps.append(json.load(open(f, encoding="utf-8")))
        except Exception:
            pass
    return snaps

def evolution_md(snaps):
    """Tables d'évolution — ou le refus explicite d'en produire.

    Défaut corrigé le 2026-08-03. Les éditions du 6 et du 9 juillet portaient des
    instantanés IDENTIQUES au chiffre près : mêmes totaux (18 429 actives,
    8 847 ventes, 9 582 locations, 2 469 condos) et **48 quartiers sur 48
    rigoureusement égaux**. Aucun scrap n'avait eu lieu entre les deux.

    Le rapport affichait donc « Δ +0.0 % » sur chaque quartier — ce qui SE LIT
    COMME UNE STABILITÉ DU MARCHÉ alors que c'est la même mesure présentée deux
    fois. Un lecteur extérieur y voyait une information ; il n'y en avait aucune.

    Une section d'évolution ne doit exister que si les données ont bougé. Quand
    elles n'ont pas bougé, il faut le DIRE, pas produire un tableau de zéros.
    """
    if len(snaps) < 2:
        return ("*Première édition snapshotée — les tables d'évolution apparaîtront "
                "automatiquement à partir de la 2e édition.*\n")

    # ─────────── artefacts de TRANSFERT, pas de marché ───────────
    # Une remontée de données locales fait entrer des milliers d'annonces d'un
    # coup. Les écarts entre éditions mélangent alors mouvement réel et effet de
    # corpus, et rien ne les distingue à la lecture. Le 2026-08-03 : 18 429 →
    # 27 380 actives, et Bang Na affichait −22,9 % sans qu'un seul prix ait bougé.
    # Les remontées sont journalisées par ops/remonter-local.py.
    avert = []
    ev = os.path.join(ROOT, "study", "evenements", "remontees.jsonl")
    if os.path.exists(ev):
        for ligne in open(ev, encoding="utf-8"):
            ligne = ligne.strip()
            if not ligne:
                continue
            try:
                e = json.loads(ligne)
            except json.JSONDecodeError:
                continue
            d = (e.get("date") or "")[:10]
            if snaps[-2]["date"] < d <= snaps[-1]["date"]:
                avert.append(e)

    a, b = snaps[-2], snaps[-1]
    if a.get("totals") == b.get("totals") and a.get("khet_stats_01") == b.get("khet_stats_01"):
        return (f"> ⚠ **Aucune évolution mesurable.** L'instantané du {b['date']} est "
                f"identique à celui du {a['date']} — mêmes totaux, mêmes valeurs sur "
                f"tous les quartiers. Aucun scrap n'a eu lieu entre les deux éditions.\n>\n"
                f"> Les tables d'évolution ne sont pas affichées : elles ne montreraient "
                f"que des « +0,0 % », ce qui se lirait à tort comme une stabilité du "
                f"marché. Relancer un cycle de scraps avant la prochaine édition.\n")

    dates = [s["date"] for s in snaps]
    out = []
    if avert:
        n = sum(e.get("annonces_transferees") or 0 for e in avert)
        out.append(
            "> ⚠ **Ces écarts ne sont pas que du marché.** "
            f"{len(avert)} remontée(s) de données locales ont eu lieu depuis "
            f"l'édition du {a['date']}, portant {n} annonces — le corpus passe de "
            f"{a['totals']['actives']} à {b['totals']['actives']} actives. "
            "Un écart de quelques pourcents relève ici du changement de composition "
            "autant que du prix.\n>\n"
            "> Ce qui reste lisible : le SENS et le CLASSEMENT relatif des quartiers. "
            "Ce qui ne l'est pas : l'ampleur. La prochaine édition, sur corpus stable, "
            "donnera la première évolution propre.\n")
    out.append(panier_constant_md(a, b))
    out.append(f"Éditions comparées : {', '.join(dates)} "
               f"(versions de config : {', '.join(str(s['config_version']) for s in snaps)}).\n")
    out.append("**Prix de vente /m² (0–1BR, double médiane) :**\n")
    header = "| Quartier | " + " | ".join(dates) + " | Δ dernière |"
    out.append(header)
    out.append("|---" * (len(dates) + 2) + "|")
    for khet in CFG["key_khets_evolution"]:
        vals = [s["khet_stats_01"].get(khet, {}).get("sale_psqm") for s in snaps]
        if not any(vals):
            continue
        delta = ""
        if vals[-1] and vals[-2]:
            delta = f"{(vals[-1] - vals[-2]) / vals[-2] * 100:+.1f} %"
        out.append(f"| {khet.replace(' District', '')} | "
                   + " | ".join(thb(v) if v else "—" for v in vals) + f" | {delta} |")
    out.append("\n**Rendement within-condo (0–1BR, %) :**\n")
    out.append(header)
    out.append("|---" * (len(dates) + 2) + "|")
    for khet in CFG["key_khets_evolution"]:
        vals = [s["khet_stats_01"].get(khet, {}).get("yield_wc") for s in snaps]
        if not any(vals):
            continue
        delta = ""
        if vals[-1] and vals[-2]:
            delta = f"{vals[-1] - vals[-2]:+.2f} pt"
        out.append(f"| {khet.replace(' District', '')} | "
                   + " | ".join(f"{v:.2f}" if v else "—" for v in vals) + f" | {delta} |")
    out.append("\n**Taux de délistage par édition (churn du stock ventes, %) :**\n")
    out.append(header)
    out.append("|---" * (len(dates) + 2) + "|")
    keys = sorted({k for s in snaps for k in s.get("tension_by_khet", {})})
    for khet in keys:
        vals = [s.get("tension_by_khet", {}).get(khet, {}).get("rate") for s in snaps]
        delta = ""
        if vals[-1] is not None and vals[-2] is not None:
            delta = f"{vals[-1] - vals[-2]:+.1f} pt"
        out.append(f"| {khet.replace(' District', '')} | "
                   + " | ".join(f"{v:.1f}" if v is not None else "—" for v in vals) + f" | {delta} |")
    return "\n".join(out) + "\n"

# ───────────────────────── exports tableaux/graphiques ─────────────────────────
def _write_csv(path, header, rows, excel_fr=False):
    """CSV. excel_fr=True → séparateur ';' + décimale ',' (double-clic Excel FR)."""
    sep = ";" if excel_fr else ","
    def fmt(v):
        if v is None:
            return ""
        if isinstance(v, float):
            s = f"{v:.2f}".rstrip("0").rstrip(".")
            return s.replace(".", ",") if excel_fr else s
        return str(v)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        f.write(sep.join(header) + "\n")
        for r in rows:
            f.write(sep.join(fmt(v) for v in r) + "\n")

def export_tables(D01, tens, snaps, demo=None):
    """Tables par quartier prêtes à grapher : CSV (std + Excel FR) + xlsx avec graphiques."""
    ddir = os.path.join(ROOT, CFG["paths"]["reports_dir"], "data")
    os.makedirs(ddir, exist_ok=True)
    tens_by_khet = {r["khet"]: r for r in tens["by_khet"]}
    demo = demo or {}

    # 1) édition courante : 1 ligne par quartier, toutes métriques
    header = ["quartier", "vente_psqm_01", "loyer_psqm_01", "rendement_wc_01",
              "immeubles_vente", "immeubles_loc", "apparies", "low_sample",
              "delistees_3j", "taux_delistage_pct",
              "population_dopa", "ventes_pour_1000_hab", "locs_pour_1000_hab"]
    cur_rows = []
    for khet, v in sorted(D01.items(), key=lambda kv: -(kv[1]["sale_psqm"] or 0)):
        t = tens_by_khet.get(khet, {})
        d = demo.get(khet, {})
        cur_rows.append([khet.replace(" District", ""), v["sale_psqm"], v["rent_psqm"],
                         v["yield_wc"], v["sale_condos"], v["rent_condos"], v["paired"],
                         1 if v["low_sample"] else 0, t.get("n_delisted"), t.get("delist_rate_pct"),
                         d.get("population"), d.get("sale_per_1000"), d.get("rent_per_1000")])
    _write_csv(os.path.join(ddir, f"khet-{TODAY}.csv"), header, cur_rows)
    _write_csv(os.path.join(ddir, f"khet-{TODAY}-excel-fr.csv"), header, cur_rows, excel_fr=True)

    # 2) séries longues (toutes éditions × tous quartiers) — format tidy pour pivots/graphiques
    sheader = ["date", "quartier", "vente_psqm_01", "loyer_psqm_01", "rendement_wc_01",
               "apparies", "taux_delistage_pct", "config_version"]
    srows = []
    for s in snaps:
        tk = s.get("tension_by_khet", {})
        for khet, v in sorted(s.get("khet_stats_01", {}).items()):
            srows.append([s["date"], khet.replace(" District", ""), v.get("sale_psqm"),
                          v.get("rent_psqm"), v.get("yield_wc"), v.get("paired"),
                          tk.get(khet, {}).get("rate"), s["config_version"]])
    _write_csv(os.path.join(ddir, "series-khet.csv"), sheader, srows)
    _write_csv(os.path.join(ddir, "series-khet-excel-fr.csv"), sheader, srows, excel_fr=True)

    # 3) xlsx avec graphiques natifs
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
    except ImportError:
        print("  (openpyxl absent → xlsx sauté ; pip install openpyxl)")
        return ddir
    wb = Workbook()

    ws = wb.active
    ws.title = "Quartiers"
    ws.append(header)
    for r in cur_rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18

    ws2 = wb.create_sheet("Séries")
    ws2.append(sheader)
    for r in srows:
        ws2.append(r)
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["B"].width = 18

    # feuilles larges (date en ligne, quartiers-clés en colonnes) pour les courbes
    keys = [k.replace(" District", "") for k in CFG["key_khets_evolution"]]
    def wide_sheet(name, metric):
        w = wb.create_sheet(name)
        w.append(["date"] + keys)
        for s in snaps:
            row = [s["date"]]
            for k in CFG["key_khets_evolution"]:
                row.append(s.get("khet_stats_01", {}).get(k, {}).get(metric))
            w.append(row)
        w.sheet_state = "hidden"
        return w
    wprix = wide_sheet("_prix", "sale_psqm")
    wyld = wide_sheet("_yield", "yield_wc")

    wg = wb.create_sheet("Graphiques")
    n = len(cur_rows)

    bar1 = BarChart(); bar1.type = "col"; bar1.title = "Prix de vente /m² (0–1BR, double médiane)"
    bar1.height, bar1.width = 10, 30
    bar1.add_data(Reference(ws, min_col=2, min_row=1, max_row=n + 1), titles_from_data=True)
    bar1.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    bar1.legend = None
    wg.add_chart(bar1, "A1")

    bar2 = BarChart(); bar2.type = "col"; bar2.title = "Rendement within-condo % (0–1BR)"
    bar2.height, bar2.width = 10, 30
    bar2.add_data(Reference(ws, min_col=4, min_row=1, max_row=n + 1), titles_from_data=True)
    bar2.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    bar2.legend = None
    wg.add_chart(bar2, "A22")

    n_ed = len(snaps)
    line1 = LineChart(); line1.title = "Évolution prix /m² — quartiers clés"
    line1.height, line1.width = 10, 30
    line1.add_data(Reference(wprix, min_col=2, max_col=len(keys) + 1, min_row=1, max_row=n_ed + 1),
                   titles_from_data=True)
    line1.set_categories(Reference(wprix, min_col=1, min_row=2, max_row=n_ed + 1))
    wg.add_chart(line1, "A43")

    line2 = LineChart(); line2.title = "Évolution rendement WC % — quartiers clés"
    line2.height, line2.width = 10, 30
    line2.add_data(Reference(wyld, min_col=2, max_col=len(keys) + 1, min_row=1, max_row=n_ed + 1),
                   titles_from_data=True)
    line2.set_categories(Reference(wyld, min_col=1, min_row=2, max_row=n_ed + 1))
    wg.add_chart(line2, "A64")

    xpath = os.path.join(ddir, f"etude-{TODAY}.xlsx")
    wb.save(xpath)
    return ddir

# ───────────────────────── rendu ─────────────────────────
def render(actives, D01, A, B, tens, snaps, db_start, demo=None, official=None):
    demo = demo or {}
    o, p, t = CFG["opportunities"], CFG["proximity"], CFG["tension"]
    n_sales = sum(1 for r in actives if r["deal_type"] == "sale")
    n_rents = sum(1 for r in actives if r["deal_type"] == "rent")
    L = []
    L.append(f"# Étude de marché Bangkok condos — édition du {TODAY}\n")
    L.append(f"*Framework figé v{CFG['config_version']} (study/config.json) — reproductible : "
             f"`python study/run_study.py` après chaque cycle de scraps complets. "
             f"Base depuis le {db_start}. Prix affichés, THB.*\n")
    L.append(f"**Corpus** : {len(actives):,} annonces actives ({n_sales:,} ventes / {n_rents:,} locations), "
             f"{len({norm_condo(r['condo_name']) for r in actives}):,} condos distincts. "
             f"Délistées analysées ({t['delisted_window_days']} j) : {tens['delisted_total']:,} "
             f"(sources : {', '.join(tens['sources'])}).\n".replace(",", " "))

    L.append("## 1. Méthode (figée)\n")
    L.append(f"Double médiane par condo (1 immeuble = 1 voix), winsorisation p{CFG['stats']['winsor_pct'][0]}–"
             f"p{CFG['stats']['winsor_pct'][1]}, strate {CFG['stats']['strata_headline']}, rendement within-condo "
             f"(≥{CFG['stats']['min_paired_condos']} immeubles appariés), bornes ventes "
             f"{thb(CFG['stats']['sale_min_thb'])}–{thb(CFG['stats']['sale_max_thb'])} THB, loyers ≤"
             f"{thb(CFG['stats']['rent_max_thb_month'])}/mois. Opportunité : décote vs médiane du même immeuble "
             f"(hors annonce, ≥{o['min_sales_in_condo_for_discount']} ventes) et/ou renta ≥{o['yield_floor_pct']} % "
             f"(≥{o['min_rents_in_condo_for_yield']} loyers) ; ⚠ si décote >{o['flag_discount_pct']} % ou renta "
             f">{o['flag_yield_pct']} %. Limites : prix affichés ; délistage ≠ vente ; "
             "comparaisons entre quartiers valides, niveaux absolus indicatifs.\n")

    dopa_note = ""
    if official and "dopa" in official:
        dopa_note = (f" Population = registre DOPA (mois bouddhiste "
                     f"{official['dopa']['yymm_buddhist']}) ; annonces/1000 hab toutes strates. "
                     "⚠ le registre sous-compte les résidents NON enregistrés (expats, locataires) — "
                     "dans les quartiers expat, lire les /1000 comme un indice de densité d'offre, "
                     "pas un taux réel ; c'est l'ÉVOLUTION de la population qui portera le signal.")
    L.append(f"## 2. État du marché (0–1BR, double médiane par condo){' ' if dopa_note else ''}\n")
    if dopa_note:
        L.append(f"*{dopa_note.strip()}*\n")
    L.append("| Quartier | Vente /m² | Loyer /m²/mois | Rendement WC | Immeubles S/R | Pop. (DOPA) | Ventes/1000 | Locs/1000 |")
    L.append("|---|---|---|---|---|---|---|---|")
    for khet, v in sorted(D01.items(), key=lambda kv: -(kv[1]["sale_psqm"] or 0)):
        if (v["sale_condos"] + v["rent_condos"]) < 10:
            continue
        ls = " *(low sample)*" if v["low_sample"] else ""
        d = demo.get(khet, {})
        L.append(f"| {khet.replace(' District', '')}{ls} | {thb(v['sale_psqm'])} | "
                 f"{v['rent_psqm'] or '—'} | {v['yield_wc'] or '—'} % | {v['sale_condos']}/{v['rent_condos']} | "
                 f"{thb(d.get('population'))} | {d.get('sale_per_1000') or '—'} | {d.get('rent_per_1000') or '—'} |")

    if official and official.get("bot", {}).get("series"):
        L.append("\n### Calibrage vs indice officiel (BOT/REIC)\n")
        L.append("| Période | Indice condo BKK | y/y |")
        L.append("|---|---|---|")
        for s in official["bot"]["series"]:
            L.append(f"| {s['period']} | {s['condo_index_bkk']} | {s.get('yoy_pct', '—')} % |")
        L.append("\n*Nos médianes = prix affichés ; l'indice = transactions financées (hédonique). "
                 "L'écart de trajectoire entre les deux mesure l'évolution de la marge de négociation. "
                 "Série à compléter à chaque publication (study/official/bot-manual.json), "
                 "automatisable avec une clé API BOT gratuite.*")

    L.append("\n## 3. Évolution entre éditions\n")
    L.append(evolution_md(snaps))

    L.append("\n## 4. Contexte & perspectives (narratif manuel — study/context.md)\n")
    L.append(open(os.path.join(ROOT, CFG["paths"]["context_md"]), encoding="utf-8").read())

    L.append(f"\n## 5. Opportunités par quartier expat (top {o['top_n_per_khet']}/quartier)\n")
    for khet, items in A.items():
        L.append(f"\n### {khet.replace(' District', '')}\n")
        for i, e in enumerate(items, 1):
            beds = "Studio" if (e["beds"] or 0) == 0 else f"{e['beds']}BR"
            sig = []
            if e["discount_pct"] is not None:
                sig.append(f"décote condo **{e['discount_pct']:+.1f} %** ({e['n_sale_condo']} ventes)")
            if e["yield_pct"] is not None:
                sig.append(f"renta est. **{e['yield_pct']:.1f} %** ({e['n_rent_condo']} loyers)")
            warn = " ⚠ *à vérifier (surface/tenure)*" if e["flag"] else ""
            area = f" · {e['area']:.0f} m²" if e["area"] else ""
            L.append(f"{i}. **{e['condo']}** — {beds}{area} · **{thb(e['price'])} THB** "
                     f"({thb(e['ppsqm'])}/m²)  \n   {' · '.join(sig)} · depuis {e['first_seen']} · "
                     f"[annonce ↗]({e['url']}) `{e['source']}`{warn}")

    L.append(f"\n## 6. École ≤{CFG['proximity']['school_max_m'] / 1000:.0f} km + métro ≤"
             f"{p['metro_max_m']} m (existant ou futur) — top {p['top_n_report']}\n")
    L.append("| # | Condo | Quartier | Bien | Prix THB | /m² | Renta | École (m) | Métro (m) |")
    L.append("|---|---|---|---|---|---|---|---|---|")
    for i, e in enumerate(B[:p["top_n_report"]], 1):
        beds = "St" if (e["beds"] or 0) == 0 else f"{e['beds']}BR"
        metro = (f"{e['metro']} ({e['metro_m']})" if e["metro_m"] is not None
                 else f"⧖ {e['future_line']} ({e['future_m']})")
        area = f" {e['area']:.0f} m²" if e["area"] else ""
        warn = " ⚠" if e["flag"] else ""
        L.append(f"| {i} | [{e['condo']}]({e['url']}){warn} | {e['khet'].replace(' District', '')} | "
                 f"{beds}{area} | {thb(e['price'])} | {thb(e['ppsqm'])} | **{e['yield_pct']:.1f} %** | "
                 f"{e['school']} ({e['school_m']}) | {metro} |")

    L.append(f"\n## 7. Tension — délistées des {t['delisted_window_days']} derniers jours\n")
    cad = t.get("scan_cadence_days", 4)
    L.append("Rappel : délistage = vendu OU retiré OU artefact de fenêtre ; comparaisons entre "
             "quartiers/typologies valides, niveaux absolus non. Vies médianes plafonnées par l'âge de la base.\n")
    L.append(f"\n> ⚠ **La vie médiane ne peut pas descendre sous {cad} jours**, qui est "
             f"l'intervalle entre deux scraps. Une valeur égale à {cad} signifie « morte "
             f"avant qu'on repasse », pas « vendue en {cad} jours » — la mesure ne résout "
             f"rien en dessous de sa propre cadence. Ces lignes sont marquées `≤`.\n>\n"
             f"> S'y ajoute la republication : un agent qui retire et repose son annonce "
             f"crée une mort et une naissance sans qu'un lot change de main. C'est ce que "
             f"les cohortes `unit_key` corrigent, et que cette section n'utilise pas encore.\n")
    L.append(f"\n**Par quartier (≥{t['min_delisted_per_khet']} délistées)** :\n")
    L.append("| Quartier | Délistées | Taux | Vie médiane (j) | Prix médian |")
    L.append("|---|---|---|---|---|")
    for r in tens["by_khet"]:
        if r["n_delisted"] >= t["min_delisted_per_khet"]:
            vie = r["median_days_active"]
            # `≤` plutôt qu'une valeur nue : le chiffre est un plancher de mesure,
            # pas une observation. Le présenter nu invite à le lire comme un délai
            # de vente réel.
            vie_txt = f"≤ {cad}" if vie is not None and vie <= cad else f"{vie}"
            L.append(f"| {r['khet'].replace(' District', '')} | {r['n_delisted']} | {r['delist_rate_pct']} % | "
                     f"{vie_txt} | {thb(r['median_price'])} |")
    L.append("\n**Par typologie (churn = délistées / (délistées + stock actif des mêmes sources))** :\n")
    L.append("| Typologie | Délistées | Stock actif | Churn |")
    L.append("|---|---|---|---|")
    for r in tens["by_typology"]:
        L.append(f"| {r['typo']} | {r['n_delisted']} | {r['active']} | **{r['churn_pct']} %** |")

    L.append("\n---\n*Étude générée automatiquement (study/run_study.py). Les fiches ⚠ exigent une "
             "vérification humaine avant toute action. Usage perso non commercial.*")
    return "\n".join(L)

# ───────────────────────── main ─────────────────────────
def main():
    print(f"▶ Étude {TODAY} (config v{CFG['config_version']})")
    actives, delisted, db_start = fetch_all()
    print(f"  {len(actives)} actives, {len(delisted)} délistées ({CFG['tension']['delisted_window_days']} j)")
    schools, metros, future = load_geo()
    condo_sales, condo_rents = build_condo_index(actives)

    D01 = khet_stats(actives, "0-1BR")
    A = opportunities(actives, condo_sales, condo_rents)
    B = school_metro(actives, condo_sales, condo_rents, schools, metros, future)
    tens = tension(actives, delisted)

    official = load_official()
    demo = per_1000(actives, official)
    if demo:
        print(f"  démographie DOPA : {len(demo)} khets croisés")

    snap_path = write_snapshot(actives, D01, tens, demo, official)
    print(f"  snapshot : {snap_path}")
    snaps = load_snapshots()

    ddir = export_tables(D01, tens, snaps, demo)
    print(f"  tables/graphiques : {ddir} (khet-{TODAY}.csv ×2, series-khet.csv ×2, etude-{TODAY}.xlsx)")

    rdir = os.path.join(ROOT, CFG["paths"]["reports_dir"])
    os.makedirs(rdir, exist_ok=True)
    rpath = os.path.join(rdir, f"etude-{TODAY}.md")
    open(rpath, "w", encoding="utf-8").write(
        render(actives, D01, A, B, tens, snaps, db_start, demo, official))
    print(f"✓ rapport : {rpath}")
    print(f"  ({len(snaps)} snapshot(s) → évolution {'active' if len(snaps) >= 2 else 'dès la 2e édition'})")

if __name__ == "__main__":
    main()
